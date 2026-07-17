"""Figure 5 (fig:validation): internal validation of the pilot Arrhenius extraction.

Two panels:
  (a) Time-window cross-validation. A linear DeltaE*(t) rate is fitted per spot on
      the early measurements (days 0-12) and used to predict the late-window
      measurements (days 25/27/31/34). Each point is a (spot, test-day) pair,
      predicted vs observed DeltaE*_ab, colour-coded by environmental arm. The
      dashed line is y = x (perfect prediction). RMSE / MAE / Q^2 (out-of-sample) are computed
      here from the recovered (definitive) measured data at run time.
  (b) Room/chamber fade-rate ratios with bootstrap 95% CIs (vermilion ~2.0,
      azurite ~1.7, malachite ~1.5; per-spot linear DeltaE*(t) slopes over the
      full 61-day window, 2000 spot-index resamples). The two arms differ in
      T, RH and illumination simultaneously, so we report the measured rate
      contrast rather than a deconvolved activation energy.

DeltaE computed directly from L*a*b* vs each spot's day-0 baseline.
Single caption in the paper; plot + axis labels only here.
Writes to the Heritage-Sciences paper figures folder (and a local copy).
"""
import os, math
import numpy as np
import openpyxl
import matplotlib as mpl
import matplotlib.pyplot as plt
_OUT_DIR = os.path.dirname(__file__)

mpl.rcParams.update({"font.family": "sans-serif", "font.size": 9, "axes.linewidth": 0.8,
                     "pdf.fonttype": 42, "ps.fonttype": 42})
INK = "#333333"; MARK = "#0072B2"
ROOM = "#0072B2"; CHAM = "#E63946"

# Dataset is distributed via figshare (see experiments/README.md); place the
# downloaded file in experiments/data/ before running.
DATA = os.path.join(os.path.dirname(__file__), "..", "data", "mogao_data_measured_recovered.xlsx")
PIGS = ["Vermilion", "Azurite", "Malachite"]
TRAIN_MAX = 12
TEST_DAYS = [25, 27, 31, 34]


def _rows(wb, sheet):
    ws = wb[sheet]; hdr = [c.value for c in ws[1]]
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def cross_validation(wb):
    """Fit linear DeltaE*(t) per spot on days<=12, predict test days; return pairs + stats."""
    base = {}
    for pig in PIGS:
        for d in _rows(wb, pig):
            if d["L*"] is None:
                continue
            key = (pig, str(d["geometry"]), str(d["arm_code"]), str(d["spot"]))
            if d["day"] == 0:
                base[key] = (d["L*"], d["a*"], d["b*"])
    series = {}
    for pig in PIGS:
        for d in _rows(wb, pig):
            key = (pig, str(d["geometry"]), str(d["arm_code"]), str(d["spot"]))
            if d["L*"] is None or key not in base:
                continue
            L0, a0, b0 = base[key]
            dE = math.sqrt((d["L*"] - L0) ** 2 + (d["a*"] - a0) ** 2 + (d["b*"] - b0) ** 2)
            series.setdefault(key, []).append((d["day"], dE))
    pred, obs, arms = [], [], []
    for key, ser in series.items():
        ser = sorted(ser)
        train = [(t, e) for t, e in ser if t <= TRAIN_MAX]
        if len(train) < 2:
            continue
        t = np.array([x[0] for x in train]); y = np.array([x[1] for x in train])
        m, c = np.polyfit(t, y, 1)
        arm = "room" if "room" in key[2] else "chamber"
        for td, oe in ser:
            if td in TEST_DAYS:
                pred.append(m * td + c); obs.append(oe); arms.append(arm)
    pred, obs = np.array(pred), np.array(obs)
    rmse = math.sqrt(np.mean((pred - obs) ** 2)); mae = np.mean(np.abs(pred - obs))
    r2 = 1 - np.sum((obs - pred) ** 2) / np.sum((obs - obs.mean()) ** 2)
    return pred, obs, arms, dict(n=len(pred), rmse=rmse, mae=mae, r2=r2)


wb = openpyxl.load_workbook(DATA, data_only=True)
pred, obs, arms, stats = cross_validation(wb)
print(f"CV: n={stats['n']} RMSE={stats['rmse']:.2f} MAE={stats['mae']:.2f} R2={stats['r2']:.1f}")

# Room/chamber fade-rate ratio per pigment, with bootstrap 95% CIs, computed here
# from mogao_data_measured_recovered.xlsx (the definitive source): per-spot linear
# DeltaE*(t) slopes on the pedestal specimens over the full 61-day window, then the
# ratio of the mean room slope to the mean chamber slope. 2000 spot-index resamples
# (seed 42). Because the two arms differ in temperature, humidity AND illumination
# (lit room vs dark chamber) simultaneously, this ratio is the primary measurable
# result; it is NOT deconvolved into an activation energy (see Methods/Results).
def _rate_ratio_ci(wb, seed=42, nboot=2000):
    base = {}
    for pig in PIGS:
        for d in _rows(wb, pig):
            if d["L*"] is None:
                continue
            key = (pig, str(d["arm_code"]), str(d["spot"]))
            if d["day"] == 0:
                base[key] = (d["L*"], d["a*"], d["b*"])
    ser = {}
    for pig in PIGS:
        for d in _rows(wb, pig):
            key = (pig, str(d["arm_code"]), str(d["spot"]))
            if d["L*"] is None or key not in base:
                continue
            L0, a0, b0 = base[key]
            dE = math.sqrt((d["L*"] - L0) ** 2 + (d["a*"] - a0) ** 2 + (d["b*"] - b0) ** 2)
            ser.setdefault(key, []).append((d["day"], dE))
    rng = np.random.default_rng(seed)
    out = []
    for pig in PIGS:
        room, cham = {}, {}
        for key, pts in ser.items():
            if key[0] != pig:
                continue
            pts = np.array(sorted(pts))
            s = np.polyfit(pts[:, 0], pts[:, 1], 1)[0]
            (room if "room" in key[1] else cham)[key[2]] = s
        rk, ck = list(room), list(cham)
        ratio = np.mean(list(room.values())) / np.mean(list(cham.values()))
        boot = []
        for _ in range(nboot):
            rr = np.mean([room[rk[i]] for i in rng.integers(0, len(rk), len(rk))])
            cc = np.mean([cham[ck[i]] for i in rng.integers(0, len(ck), len(ck))])
            if cc > 0:
                boot.append(rr / cc)
        lo, hi = np.percentile(boot, [2.5, 97.5])
        out.append((pig, round(ratio, 1), round(lo, 1), round(hi, 1)))
    return out

ci = _rate_ratio_ci(wb)
print("rate ratios:", ci)

fig, (axA, axB) = plt.subplots(1, 2, figsize=(11, 4.2))

# --- Panel (a): cross-validation scatter ---
arms = np.array(arms)
for arm, col, lbl in [("room", ROOM, "23\u00b0C / 40% RH (room)"),
                      ("chamber", CHAM, "40\u00b0C / 10% RH (chamber)")]:
    mask = arms == arm
    axA.scatter(obs[mask], pred[mask], s=26, c=col, alpha=0.7, edgecolors="white",
                linewidths=0.5, label=lbl, zorder=3)
lim = max(obs.max(), pred.max()) * 1.08
axA.plot([-1, lim], [-1, lim], ls="--", color="#999999", lw=1.1, zorder=1, label="perfect prediction")
axA.set_xlim(-0.3, lim); axA.set_ylim(min(pred.min() * 1.1, -0.3), lim)
axA.set_xlabel(r"observed $\Delta E^*_{ab}$ (days 25--34)")
axA.set_ylabel(r"predicted $\Delta E^*_{ab}$ (from days 0--12 rate)")
axA.set_title("(a) Time-window cross-validation", loc="left", fontsize=10, fontweight="bold")
axA.legend(fontsize=7.5, frameon=False, loc="upper left")
axA.text(0.97, 0.05,
         f"RMSE $= {stats['rmse']:.2f}$\nMAE $= {stats['mae']:.2f}$\n$Q^2 = {stats['r2']:.1f}$  ($n={stats['n']}$)",
         transform=axA.transAxes, ha="right", va="bottom", fontsize=8.5, color=INK,
         bbox=dict(boxstyle="round,pad=0.35", fc="white", ec="#CCCCCC", lw=0.8))
axA.grid(True, color="#EEEEEE", lw=0.6, zorder=0)
for sp in ("top", "right"):
    axA.spines[sp].set_visible(False)

# --- Panel (b): bootstrap Ea CIs ---
y = list(range(len(ci)))[::-1]
for (name, est, lo, hi), yi in zip(ci, y):
    axB.plot([lo, hi], [yi, yi], color=MARK, lw=2.2, solid_capstyle="round", zorder=2)
    axB.plot([lo, lo], [yi - 0.09, yi + 0.09], color=MARK, lw=1.4)
    axB.plot([hi, hi], [yi - 0.09, yi + 0.09], color=MARK, lw=1.4)
    axB.plot(est, yi, "o", color=MARK, mec="white", mew=0.8, ms=7, zorder=3)
    axB.annotate(rf"{est}$\times$  [{lo}, {hi}]", (hi, yi), xytext=(6, 0), textcoords="offset points",
                 va="center", fontsize=8.5, color=INK)
axB.set_yticks(y); axB.set_yticklabels([d[0] for d in ci], fontsize=10)
axB.axvline(1, color="#BBBBBB", lw=0.9, ls="--", zorder=1)  # ratio = 1: arms age equally
axB.set_ylim(-0.6, len(ci) - 0.4); axB.set_xlim(0, 9)
axB.set_xlabel("room/chamber fade-rate ratio")
axB.set_xticks([0, 1, 2, 4, 6, 8])
axB.set_title("(b) Room/chamber fade-rate ratio (95% CI)", loc="left", fontsize=10, fontweight="bold")
axB.grid(True, axis="x", color="#E6E6E6", lw=0.6, zorder=0)
for sp in ("top", "right", "left"):
    axB.spines[sp].set_visible(False)
axB.tick_params(length=0, colors=INK)

fig.subplots_adjust(left=0.09, right=0.97, bottom=0.14, top=0.92, wspace=0.28)

_fig_dir = os.path.join(os.path.dirname(__file__), "..", "..", "Heritage-Sciences", "figures")
targets = [os.path.join(_OUT_DIR, "validation_predicted_vs_observed.png")]
if os.path.isdir(_fig_dir):
    targets.append(os.path.join(_fig_dir, "validation_predicted_vs_observed.png"))
for path in targets:
    fig.savefig(path, dpi=300, bbox_inches="tight")
print("Saved two-panel validation figure.")
